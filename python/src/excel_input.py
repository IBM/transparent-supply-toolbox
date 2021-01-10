#!/usr/bin/python

from analib import api_inbound as api_inbound
from analib import datagen as datagen
from openpyxl import load_workbook

import utils
import organization

# class for excel input.
# input is an adapter. adapter should support specific functions

class ExcelInput:

    filename = ''
    adapter = None
    specific_rows = []

    def __init__(self, filename, adapter):
        self.filename = filename
        self.adapter = adapter
        return

    # print single excel row as a table
    # the function below needs to be worked on
    cIndex = 0
    def __debug_print_for_single_xlrow(self, row):
        cIndex = 0
        data_points = {}
        def add_value(key_name):
            global cIndex
            if row[cIndex] == None:
                data_points[key_name] = ''
            else:
                print (row[cIndex])
                data_points[key_name] = row[cIndex]
            cIndex += 1
            return cIndex

        first_row = ['Field Name', 'Value']

        # this should be the order in excel sheet
        add_value('Type')
        add_value('Time')
        add_value('ID')
        add_value('Name')    
        add_value('Input Type')
        add_value('Input')
        add_value('Output Type')
        add_value('Output')
        add_value('BizStep')
        add_value('BizLoc')    
        add_value('Src')
        add_value('Dest')
        add_value('BizTxn')
        utils.d_pprint(0, "Event: ", data_points)
        utils.print_table_from_dict(first_row, data_points, alternate_color=True)


    def _check_row_length(self, row_no, len_d_row, expected_params):
        if len_d_row != expected_params:
            print ("Row No = ", row_no)
            print ("Returned Length by adapter: len(d_rows): ", len_d_row, ". Expected # params: ", expected_params)
            utils.quit("Params in Row do not match.")


    # takes in the argument returned from load_workbook
    # b_rows stands for "business format rows" - e.g., rows populated by a business user
    # d_rows stands for "desired format rows"  - rows that are understood by underlying classes
    # adapter convers from b_rows to d_rows.  
    def parse_and_push_worksheet(self, org_params, b_rows, data_type, inbound_api):

        # criteria that determines if a row is to be pushed
        def row_to_be_skipped(row_no):
            utils.d_print(6, "Row: ", row_no, "specific_rows:", self.specific_rows)
            ret_val = False
            # skip the first row; it is the names of various columns; we start with row_no 1.
            if row_no == 0:
                ret_val = True
            elif self.specific_rows:  # we check if this is not None
                if str(row_no) not in self.specific_rows:
                    ret_val = True
            utils.d_print(6, "Row: ", row_no, "Skip?:", ret_val)
            return ret_val

        try:
            b_row = {}            
            row_no = 0
            for row in b_rows.iter_rows(values_only=True):
                # if this row is not to be pushed based on certain criteria, increment and continue
                if row_to_be_skipped(row_no):
                    row_no += 1 
                    continue

                 # this is a row to be pushed.
                index = 0
                ## we check for the first column.
                if row[0] is not None:
                    utils.d_print (2, "-------> ", row)
                    for cell_value in row:
                        # if cell_value is not None: 
                        # the value of index should match the values of F_PARAMS in datagen
                        # for Strings, we have a flag to control if we want to encode as unicode bytes
                        unicode_encode = False
                        if unicode_encode and isinstance(cell_value, str):
                            b_row[index] = cell_value.encode('utf-8')
                        else:
                            b_row[index] = cell_value
                        index += 1
                    utils.d_pprint(3, "Business Format Row: ", b_row)

                    # we now have read the "business row"
                    # this is where we convert the row we just read using the adapter to the row
                    # that we want.
                    if data_type == 'Facility':
                        d_row = self.adapter.convert_facility_row(b_row)
                        self._check_row_length(row_no+1, len(d_row), datagen.N_FACILITY_PARAMS)
                        x = datagen.Facility(org_params, d_row)
                    elif data_type == 'Product':
                        d_row = self.adapter.convert_product_row(b_row)
                        self._check_row_length(row_no+1, len(d_row), datagen.N_PRODUCT_PARAMS)
                        x = datagen.Product(org_params, d_row)
                    elif data_type == 'Event':
                        d_row = self.adapter.convert_event_row(b_row)
                        self._check_row_length(row_no+1, len(d_row), datagen.N_EVENT_PARAMS)
                        x = datagen.Event(org_params, d_row)
                    elif data_type == 'Payload':
                        d_row = self.adapter.convert_payload_row(b_row)
                        self._check_row_length(row_no+1, len(d_row), datagen.N_PAYLOAD_PARAMS)
                        x = datagen.Payload(org_params, d_row)
                    else:
                        utils.quit('Unknown Data Type!')

                    # retrieve the XML and push it.
                    xml = x.get_xml()
                    utils.d_print(5, "xml: ", xml)
                    inbound_api.push_xml(xml)

                row_no += 1 

                print ((row_no - 1), " ", data_type, " elements pushed.")

            # there could be additional payloads left in the adapter
            # this is the case where there is no explicit payload tab.
            if self.adapter.get_built_count('Payloads') > 0:
                a_payloads = self.adapter.get_built_data('Payloads')
                payload_index = 0
                for p_payload in a_payloads:
                    self._check_row_length(payload_index+1, len(p_payload), datagen.N_PAYLOAD_PARAMS)
                    x = datagen.Payload(org_params, p_payload)
                    xml = x.get_xml()
                    utils.d_print(5, "xml: ", xml)
                    if not utils.global_args.simulate:
                        inbound_api.push_xml(xml)
        except:
            utils.print_exception(1)


    # main function that parses and pushes the file to BTS
    def push_to_bts(self):
        company_prefixes = []
        inbound_api = None

        try:
            # 1. load the workbook
            wb = load_workbook(filename = self.filename)

            # 2. set up the inbound API
            org = organization.Organization("Name", utils.global_args.client)
            org_id = org.get_id()
            inbound_api = api_inbound.APIInbound(5, org_id, utils.global_args.env, utils.global_args.header_entitled_org)

            # get company prefixes, head quarters GLN, etc.
            org_params = org.get_properties()
            # if we are asked to use a different HQ GLN, use it.
            if utils.global_args.hq_gln:
                org_params['hq_gln'] = utils.global_args.hq_gln

            # 3. start parsing the excel sheet
            #     a. first gather global params
            sheet_names = wb.sheetnames
            utils.d_print(2, "Sheet Names in the Excel Sheet: ", sheet_names)


            # if arguments indicate specific rows to be pushed, push the same.
            if utils.global_args.specific_rows:
                self.specific_rows = utils.global_args.specific_rows.split(',')
            else:
                self.specific_rows = None

            if 'if' in utils.global_args.isheets:
                # b. next push if there are any facilities
                facilities = wb['Facilities']
                self.parse_and_push_worksheet(org_params, facilities, 'Facility', inbound_api)

            if 'ip' in utils.global_args.isheets:
                # c. next push any products   
                products = wb['Products']
                self.parse_and_push_worksheet(org_params, products, 'Product',inbound_api)

            if 'ie' in utils.global_args.isheets:
                # d. next push any events
                events = wb['Events']
                self.parse_and_push_worksheet(org_params, events, 'Event',inbound_api)

            if 'ipl' in utils.global_args.isheets:
                # e. next push any events
                payloads = wb['Payloads']
                self.parse_and_push_worksheet(org_params, payloads, 'Payload',inbound_api)
        except:
            utils.print_exception(1)

            return

